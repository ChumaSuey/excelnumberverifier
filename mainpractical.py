import openpyxl

# Set the full path to the Excel file
filename = 'C:\\Users\\luism\\Downloads\\TRAMITES RECIBIDOS 30-11-2023.xlsx'

# Read the Excel file
wb = openpyxl.load_workbook(filename)

#UNUSED SECTION

# Check if the first worksheet exists
#if 1 in wb.sheetnames:
    # Retrieve the first worksheet
    #sheet_name = wb[1]
#else:
    # If the first worksheet doesn't exist, prompt the user to input the sheet name
    #sheet_name = input("Enter the sheet name: ")

# Get the sheet
#sheet = wb[sheet_name]

#UNUSED SECTION END

sheet = wb["1"]

# Search for the number
num_found = False
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == '07247-2023':
            num_found = True
            break

# Check if the number was found
if num_found:
    print('The number 07247-2023 is present in the Excel file {}.'.format(filename))
else:
    print('The number 07247-2023 is not present in the Excel file {}.'.format(filename))